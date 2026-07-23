#!/usr/bin/env python3
"""Importa fichas CSV de medicina estetica a historial_clinico.

Cada paciente del CSV (separado por cabecera / ';') genera UNA consulta por dia
con fecha distinta. Se intenta asociar cada consulta a una cita de agenda.

Por defecto ejecuta dry-run. Para escribir: --apply.
"""

from __future__ import annotations

import argparse
import difflib
import hashlib
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from legacy_company import DEFAULT_COMPANY_ID, MEDICINA_COMPANY_ID

# Clientes de estas fichas estan en Lipoout (María del Mar). Medicina company tiene 0 customers.
DEFAULT_CSV = Path(r"C:\Users\Lipoout\Desktop\Medicina Estética\Fichas medicina.csv")
DEFAULT_IMPORT_COMPANY_ID = DEFAULT_COMPANY_ID
SOURCE_LABEL = "Fichas medicina.csv"
MAIN_SOURCE_PREFIX = "medicina_estetica_csv_v2"
VALID_YEAR_MIN = 2020
VALID_YEAR_MAX = 2027
DATE_RE = re.compile(r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})(?!\d)")
PATIENT_RE = re.compile(r"^\s*Paciente\s+(?:de\s+)?medicina\s+est[eé]tica\s*:\s*(.*)$", re.I)
AGE_RE = re.compile(r"(\d{1,3})")


def _require_psycopg2():
    try:
        import psycopg2
        import psycopg2.extras
        from psycopg2 import errors

        return psycopg2, psycopg2.extras, errors
    except Exception as exc:  # pragma: no cover
        print(f"No se pudo importar psycopg2: {exc}", file=sys.stderr)
        sys.exit(2)


@dataclass
class ParsedDate:
    raw: str
    ymd: str | None
    valid: bool
    line_index: int
    start: int
    end: int


@dataclass
class ParsedVisit:
    """Una consulta de un dia concreto."""

    line: int
    name: str
    normalized_name: str
    source_key: str
    fecha: str | None
    raw_fecha: str | None
    edad: str
    antecedentes: str
    motivo: str
    tratamiento: str
    issues: list[str]
    is_first: bool


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def repair_text(text: str) -> str:
    if "Ã" in text or "Â" in text:
        try:
            fixed = text.encode("latin-1").decode("utf-8")
            if fixed.count("\ufffd") <= text.count("\ufffd"):
                text = fixed
        except UnicodeError:
            pass
    return "".join(ch for ch in text if ch == "\n" or ch == "\t" or ord(ch) >= 32)


def read_text_guess(path: Path) -> str:
    raw = path.read_bytes()
    candidates = [repair_text(raw.decode(enc, errors="replace")) for enc in ("utf-8-sig", "cp1252", "latin-1")]
    return min(candidates, key=lambda text: (text.count("\ufffd"), text.count("Ã")))


def normalize_name(value: str) -> str:
    value = repair_text(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^0-9A-Za-zñÑ]+", " ", value).lower().strip()
    return re.sub(r"\s+", " ", value)


def clean_line(value: str) -> str:
    return value.strip().strip(";").strip()


def clean_patient_name(value: str) -> str:
    value = clean_line(value)
    value = value.rstrip(".").strip()
    value = DATE_RE.sub(" ", value)
    value = re.sub(r"\s+", " ", value).strip(" .;-")
    return value


def parse_date_match(match: re.Match[str], line_index: int) -> ParsedDate:
    day = int(match.group(1))
    month = int(match.group(2))
    year_raw = int(match.group(3))
    year = 2000 + year_raw if year_raw < 100 else year_raw
    # Typo frecuente en fichas: 2925 → 2025
    if year > VALID_YEAR_MAX and 2900 <= year <= 2999:
        year = 2000 + (year % 100)
    try:
        parsed = date(year, month, day)
    except ValueError:
        return ParsedDate(match.group(0), None, False, line_index, match.start(), match.end())
    valid = VALID_YEAR_MIN <= parsed.year <= VALID_YEAR_MAX
    return ParsedDate(
        match.group(0),
        parsed.isoformat() if valid else None,
        valid,
        line_index,
        match.start(),
        match.end(),
    )


def all_dates(lines: list[str]) -> list[ParsedDate]:
    found: list[ParsedDate] = []
    for idx, line in enumerate(lines):
        for match in DATE_RE.finditer(line):
            found.append(parse_date_match(match, idx))
    return found


def infer_name_from_block(lines: list[str]) -> str:
    for line in lines[:8]:
        if ":" not in line:
            continue
        label, rest = line.split(":", 1)
        if normalize_name(label) != "nombre y fecha":
            continue
        return clean_line(DATE_RE.split(rest, maxsplit=1)[0])
    return ""


def split_blocks(text: str) -> list[tuple[int, int, list[str]]]:
    """Separa pacientes por cabecera. El ';' cierra el bloque previo en la practica."""
    lines = text.splitlines()
    starts: list[int] = []
    for idx, line in enumerate(lines):
        if PATIENT_RE.match(line):
            starts.append(idx)
    blocks: list[tuple[int, int, list[str]]] = []
    for pos, start in enumerate(starts):
        end = starts[pos + 1] if pos + 1 < len(starts) else len(lines)
        blocks.append((start, end, lines[start:end]))
    return blocks


def parse_section_fields(lines: list[str]) -> tuple[str, str, str, str]:
    """Devuelve (edad, antecedentes, motivo, tratamiento) de un fragmento."""
    edad_parts: list[str] = []
    antecedentes_parts: list[str] = []
    motivo_parts: list[str] = []
    tratamiento_parts: list[str] = []
    section = "antecedentes"

    for raw in lines:
        line = clean_line(raw)
        if not line:
            continue
        if PATIENT_RE.match(line):
            continue

        lower = normalize_name(line)
        if lower.startswith("nombre y fecha"):
            # Quitar fecha del valor residual si aparece en la misma linea
            continue
        if lower.startswith("edad"):
            value = clean_line(line.split(":", 1)[1] if ":" in line else line)
            if value:
                edad_parts.append(value)
            continue
        if lower.startswith("ap ") or lower == "ap" or lower.startswith("ap:"):
            section = "antecedentes"
            value = clean_line(line.split(":", 1)[1] if ":" in line else "")
            if value:
                antecedentes_parts.append(value)
            continue
        if lower.startswith("motivo de consulta") or lower.startswith("motivo"):
            section = "motivo"
            value = clean_line(line.split(":", 1)[1] if ":" in line else "")
            if value:
                motivo_parts.append(value)
            continue
        if lower.startswith("tratamiento"):
            section = "tratamiento"
            value = clean_line(line.split(":", 1)[1] if ":" in line else "")
            if value:
                tratamiento_parts.append(value)
            continue
        if lower.startswith("me ") or lower == "me" or lower.startswith("me:"):
            # Medicina estetica previa → va a antecedentes
            section = "antecedentes"
            value = clean_line(line.split(":", 1)[1] if ":" in line else line)
            if value:
                antecedentes_parts.append(value)
            continue

        if section == "motivo":
            motivo_parts.append(line)
        elif section == "tratamiento":
            tratamiento_parts.append(line)
        else:
            antecedentes_parts.append(line)

    antecedentes = "\n".join(antecedentes_parts).strip()
    if edad_parts and "edad:" not in normalize_name(antecedentes):
        antecedentes = (f"Edad: {' '.join(edad_parts)}\n{antecedentes}").strip()

    return (
        " ".join(edad_parts).strip(),
        antecedentes,
        "\n".join(motivo_parts).strip(),
        "\n".join(tratamiento_parts).strip(),
    )


def source_key_for(name: str, fecha: str | None, visit_index: int) -> str:
    base = f"{normalize_name(name)}|{fecha or ''}|v{visit_index}"
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]
    return f"{MAIN_SOURCE_PREFIX}:{digest}"


def extract_body_after_date(line: str, parsed_date: ParsedDate) -> str:
    return clean_line(line[parsed_date.end :].lstrip(":").strip())


def segment_text_for_date(lines: list[str], parsed_date: ParsedDate, end_line: int) -> str:
    """Texto de la visita: linea de la fecha (completa si la fecha no esta al inicio) + siguientes."""
    line = lines[parsed_date.line_index]
    prefix = line[: parsed_date.start].strip()
    if len(prefix) <= 2:
        first = extract_body_after_date(line, parsed_date)
    else:
        first = clean_line(line)
    parts = [first] if first else []
    parts.extend(clean_line(item) for item in lines[parsed_date.line_index + 1 : end_line])
    return "\n".join(p for p in parts if p).strip()


def parse_followup_body(text: str) -> tuple[str, str]:
    """Interpreta texto de visita sucesiva → (motivo, tratamiento)."""
    text = text.strip().lstrip(":").strip()
    if not text:
        return "Revisión", ""

    _edad, ap, motivo, tratamiento = parse_section_fields(text.splitlines())
    if motivo or tratamiento:
        trat = tratamiento
        if ap and ap not in (trat or ""):
            trat = f"{trat}\n{ap}".strip() if trat else ap
        return (motivo or "Revisión", trat or text)

    return ("Revisión", text)


def parse_patient_visits(start: int, end: int, lines: list[str]) -> list[ParsedVisit]:
    match = PATIENT_RE.match(lines[0])
    name = clean_patient_name(match.group(1) if match else "")
    if not name:
        name = clean_patient_name(infer_name_from_block(lines))
    name = name.rstrip(".").strip()

    dates = all_dates(lines)
    base_issues: list[str] = []
    if not name:
        base_issues.append("missing_customer_name")
    if not dates:
        base_issues.append("missing_exam_date")

    for parsed_date in dates:
        if not parsed_date.valid:
            base_issues.append(f"invalid_date:{parsed_date.raw}")

    valid_dates = [d for d in dates if d.valid and d.ymd]
    if not valid_dates:
        return [
            ParsedVisit(
                line=start + 1,
                name=name,
                normalized_name=normalize_name(name),
                source_key=source_key_for(name, None, 0),
                fecha=None,
                raw_fecha=dates[0].raw if dates else None,
                edad="",
                antecedentes="",
                motivo="",
                tratamiento="",
                issues=base_issues or ["missing_exam_date"],
                is_first=True,
            )
        ]

    # Primera ocurrencia de cada ymd (orden de aparicion)
    segments: list[tuple[ParsedDate, int]] = []
    seen_ymd: set[str] = set()
    for parsed_date in valid_dates:
        assert parsed_date.ymd
        if parsed_date.ymd in seen_ymd:
            continue
        seen_ymd.add(parsed_date.ymd)
        segments.append((parsed_date, parsed_date.line_index))

    first_date_line = segments[0][1]
    # Cabecera clinica: todo lo anterior a la primera fecha del bloque
    edad, antecedentes, motivo, tratamiento_header = parse_section_fields(lines[:first_date_line])

    visits: list[ParsedVisit] = []
    for visit_index, (parsed_date, seg_start) in enumerate(segments):
        # Fin del segmento = inicio de la siguiente fecha distinta
        end_line = len(lines)
        for later_date, later_start in segments[visit_index + 1 :]:
            end_line = later_start
            break

        # Fusionar textos de todas las apariciones del mismo dia
        bodies: list[str] = []
        for d in valid_dates:
            if d.ymd != parsed_date.ymd:
                continue
            nxt = len(lines)
            for later in valid_dates:
                if later.line_index > d.line_index and later.ymd != d.ymd:
                    nxt = later.line_index
                    break
            body = segment_text_for_date(lines, d, min(nxt, end_line if d.line_index == seg_start else nxt))
            if body:
                bodies.append(body)
        merged_body = "\n".join(bodies).strip()

        issues = [i for i in base_issues if not i.startswith("invalid_date")]
        if visit_index == 0:
            visit_motivo = motivo or "Consulta medicina estética"
            visit_tto = merged_body or tratamiento_header
            if tratamiento_header and merged_body and tratamiento_header not in merged_body:
                visit_tto = f"{tratamiento_header}\n{merged_body}".strip()
            visit_ap = antecedentes
        else:
            visit_motivo, visit_tto = parse_followup_body(merged_body)
            visit_ap = antecedentes
            if not visit_tto:
                issues.append("empty_followup_body")

        visits.append(
            ParsedVisit(
                line=start + parsed_date.line_index + 1,
                name=name,
                normalized_name=normalize_name(name),
                source_key=source_key_for(name, parsed_date.ymd, visit_index),
                fecha=parsed_date.ymd,
                raw_fecha=parsed_date.raw,
                edad=edad,
                antecedentes=visit_ap,
                motivo=visit_motivo,
                tratamiento=visit_tto,
                issues=issues,
                is_first=visit_index == 0,
            )
        )

    return visits


def dedupe_visits(visits: list[ParsedVisit]) -> list[ParsedVisit]:
    """El CSV puede repetir el mismo paciente; una consulta por (cliente, dia)."""
    best: dict[tuple[str, str | None], ParsedVisit] = {}
    for visit in visits:
        key = (visit.normalized_name, visit.fecha)
        prev = best.get(key)
        score = len(visit.tratamiento or "") + len(visit.motivo or "") + len(visit.antecedentes or "")
        if not prev:
            best[key] = visit
            continue
        prev_score = len(prev.tratamiento or "") + len(prev.motivo or "") + len(prev.antecedentes or "")
        if score > prev_score or (score == prev_score and visit.line > prev.line):
            # Mantener is_first si alguna version lo era
            if prev.is_first and not visit.is_first:
                visit = ParsedVisit(
                    **{
                        **visit.__dict__,
                        "is_first": True,
                        "antecedentes": visit.antecedentes or prev.antecedentes,
                        "edad": visit.edad or prev.edad,
                        "motivo": visit.motivo if visit.motivo != "Revisión" else prev.motivo,
                    }
                )
            best[key] = visit
        elif prev.is_first is False and visit.is_first:
            best[key] = ParsedVisit(
                **{
                    **prev.__dict__,
                    "is_first": True,
                    "antecedentes": prev.antecedentes or visit.antecedentes,
                    "edad": prev.edad or visit.edad,
                    "motivo": prev.motivo if prev.motivo != "Revisión" else visit.motivo,
                }
            )
    # Recalcular is_first por paciente (fecha minima)
    by_patient: dict[str, list[ParsedVisit]] = {}
    for visit in best.values():
        by_patient.setdefault(visit.normalized_name, []).append(visit)
    result: list[ParsedVisit] = []
    for group in by_patient.values():
        group_sorted = sorted(group, key=lambda v: (v.fecha or "9999", v.line))
        for idx, visit in enumerate(group_sorted):
            result.append(
                ParsedVisit(
                    **{
                        **visit.__dict__,
                        "is_first": idx == 0,
                        "source_key": source_key_for(visit.name, visit.fecha, idx),
                    }
                )
            )
    return result


def parse_csv(path: Path) -> list[ParsedVisit]:
    text = read_text_guess(path)
    visits: list[ParsedVisit] = []
    for start, end, lines in split_blocks(text):
        visits.extend(parse_patient_visits(start, end, lines))
    return dedupe_visits(visits)


def norm_code(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.lstrip("0") or "0"


def load_customers(
    cur: Any, company_id: str
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    cur.execute(
        """
        SELECT id::text, company_id::text, name, legacy_codcli, birth_date::text AS birth_date
        FROM public.customers
        WHERE company_id = %s
        """,
        (company_id,),
    )
    rows = list(cur.fetchall())
    by_name: dict[str, list[dict[str, Any]]] = {}
    by_code: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(normalize_name(row["name"] or ""), []).append(row)
        code = norm_code(row.get("legacy_codcli"))
        if code:
            by_code.setdefault(code, []).append(row)
    return by_name, by_code, rows


def load_assignments(path: Path) -> dict[int, str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise SystemExit(f"Se necesita openpyxl para leer asignaciones: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    try:
        line_idx = headers.index("Linea CSV")
    except ValueError:
        raise SystemExit("El Excel de asignaciones no tiene columna 'Linea CSV'.")
    code_idx = next((i for i, h in enumerate(headers) if h.upper().startswith("CODIGO CLIENTE")), None)
    if code_idx is None:
        raise SystemExit("El Excel de asignaciones no tiene columna 'CODIGO CLIENTE'.")

    assignments: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if line_idx >= len(row) or code_idx >= len(row):
            continue
        raw_line = row[line_idx]
        raw_code = row[code_idx]
        code = norm_code(raw_code)
        if raw_line is None or not code:
            continue
        try:
            line_no = int(str(raw_line).strip())
        except ValueError:
            continue
        assignments[line_no] = code
    return assignments


def create_customer(cur: Any, company_id: str, name: str) -> dict[str, Any]:
    cur.execute(
        """
        INSERT INTO public.customers (company_id, name)
        VALUES (%s, %s)
        RETURNING id::text, company_id::text, name, legacy_codcli, birth_date::text AS birth_date
        """,
        (company_id, name),
    )
    row = dict(cur.fetchone())
    return row


def find_customer(
    visit: ParsedVisit,
    by_name: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, list[dict[str, Any]], str | None]:
    exact = by_name.get(visit.normalized_name, [])
    if len(exact) == 1:
        return exact[0], [], None
    if len(exact) > 1:
        return None, exact, "ambiguous_customer"

    # Fuzzy: si el mejor candidato destaca, aceptarlo
    scored: list[tuple[float, dict[str, Any]]] = []
    for key in by_name.keys():
        ratio = difflib.SequenceMatcher(None, visit.normalized_name, key).ratio()
        if ratio >= 0.78:
            for row in by_name[key]:
                scored.append((ratio, row))
    scored.sort(key=lambda item: item[0], reverse=True)
    if scored:
        best_ratio, best = scored[0]
        second_ratio = scored[1][0] if len(scored) > 1 else 0.0
        # Evitar empates entre personas distintas
        same_best = [row for ratio, row in scored if abs(ratio - best_ratio) < 0.001]
        if best_ratio >= 0.86 and (best_ratio - second_ratio >= 0.03 or len({r["id"] for r in same_best}) == 1):
            return best, [row for _, row in scored[:5]], None
        candidates = []
        seen_ids: set[str] = set()
        for _ratio, row in scored[:8]:
            if row["id"] in seen_ids:
                continue
            seen_ids.add(row["id"])
            candidates.append(row)
        return None, candidates, "customer_not_found"

    return None, [], "customer_not_found"


def find_appointment(
    cur: Any,
    company_id: str,
    customer: dict[str, Any],
    fecha: str | None,
) -> tuple[str | None, list[dict[str, Any]], str | None]:
    _, _, errors = _require_psycopg2()
    if not fecha:
        return None, [], "missing_date"
    params = {
        "company_id": company_id,
        "customer_id": customer["id"],
        "legacy_codcli": (customer.get("legacy_codcli") or "").strip(),
        "name": customer["name"],
        "fecha": fecha,
    }
    try:
        cur.execute(
            """
            SELECT id::text, customer_id::text, client_name, legacy_codcli,
                   appointment_date::text AS appointment_date,
                   start_time::text AS start_time,
                   status
            FROM public.agenda_appointments
            WHERE company_id = %(company_id)s
              AND (
                customer_id = %(customer_id)s
                OR (NULLIF(%(legacy_codcli)s, '') IS NOT NULL AND legacy_codcli = %(legacy_codcli)s)
                OR lower(client_name) = lower(%(name)s)
              )
              AND (
                appointment_date = %(fecha)s::date
                OR left(start_time::text, 10) = %(fecha)s
              )
            ORDER BY start_time NULLS LAST, id
            """,
            params,
        )
    except errors.UndefinedColumn:
        cur.connection.rollback()
        cur.execute(
            """
            SELECT id::text, customer_id::text, client_name, legacy_codcli,
                   NULL::text AS appointment_date,
                   start_time::text AS start_time,
                   status
            FROM public.agenda_appointments
            WHERE company_id = %(company_id)s
              AND (
                customer_id = %(customer_id)s
                OR (NULLIF(%(legacy_codcli)s, '') IS NOT NULL AND legacy_codcli = %(legacy_codcli)s)
                OR lower(client_name) = lower(%(name)s)
              )
              AND left(start_time::text, 10) = %(fecha)s
            ORDER BY start_time NULLS LAST, id
            """,
            params,
        )
    rows = list(cur.fetchall())
    if len(rows) == 1:
        return rows[0]["id"], rows, None
    if len(rows) > 1:
        # Preferir la primera no cancelada si hay varias
        active = [r for r in rows if str(r.get("status") or "").lower() not in {"cancelled", "canceled", "anulada"}]
        pick = active[0] if len(active) == 1 else (active[0] if active else rows[0])
        if len(active) > 1 or (not active and len(rows) > 1):
            return pick["id"], rows, "ambiguous_appointment_picked_first"
        return pick["id"], rows, None
    return None, [], "appointment_not_found"


def appointment_already_linked(cur: Any, appointment_id: str, current_history_id: str | None) -> bool:
    cur.execute(
        """
        SELECT id::text
        FROM public.historial_clinico
        WHERE appointment_id = %s
          AND (%s IS NULL OR id <> %s::uuid)
        LIMIT 1
        """,
        (appointment_id, current_history_id, current_history_id),
    )
    return bool(cur.fetchone())


def find_existing_history(cur: Any, customer_id: str, source_key: str, fecha: str | None = None) -> dict[str, Any] | None:
    """Busca por import_key exacto; si no, por misma fecha de import medicina (evita v1/v2 dup)."""
    cur.execute(
        """
        SELECT id::text, appointment_id::text
        FROM public.historial_clinico
        WHERE customer_id = %s
          AND observaciones LIKE %s
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (customer_id, f"%import_key={source_key}%"),
    )
    row = cur.fetchone()
    if row:
        return dict(row)
    if not fecha:
        return None
    cur.execute(
        """
        SELECT id::text, appointment_id::text
        FROM public.historial_clinico
        WHERE customer_id = %s
          AND fecha = %s::date
          AND observaciones LIKE %s
        ORDER BY
          CASE WHEN observaciones LIKE %s THEN 0 ELSE 1 END,
          created_at DESC NULLS LAST
        LIMIT 1
        """,
        (
            customer_id,
            fecha,
            "%Fichas medicina.csv%",
            "%medicina_estetica_csv_v2:%",
        ),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def maybe_set_birth_date(cur: Any, customer: dict[str, Any], edad: str, ref_fecha: str | None) -> bool:
    """Si el cliente no tiene birth_date y hay edad, estima año de nacimiento."""
    if customer.get("birth_date"):
        return False
    if not edad or not ref_fecha:
        return False
    m = AGE_RE.search(edad)
    if not m:
        return False
    age = int(m.group(1))
    if age < 1 or age > 120:
        return False
    try:
        ref = date.fromisoformat(ref_fecha)
    except ValueError:
        return False
    approx = date(ref.year - age, 1, 1)
    cur.execute(
        """
        UPDATE public.customers
        SET birth_date = %s
        WHERE id = %s::uuid
          AND birth_date IS NULL
        """,
        (approx.isoformat(), customer["id"]),
    )
    customer["birth_date"] = approx.isoformat()
    return cur.rowcount > 0


def upsert_visit(
    cur: Any,
    visit: ParsedVisit,
    company_id: str,
    customer: dict[str, Any],
    appointment_id: str | None,
) -> str:
    existing = find_existing_history(cur, customer["id"], visit.source_key, visit.fecha)
    observaciones = f"Importado de {SOURCE_LABEL}; import_key={visit.source_key}"
    payload = {
        "customer_id": customer["id"],
        "company_id": company_id,
        "fecha": visit.fecha,
        "appointment_id": appointment_id,
        "tipo": "consulta",
        "titulo": (visit.motivo or "Consulta medicina estética")[:200],
        "descripcion": visit.antecedentes or None,
        "antecedentes_personales": visit.antecedentes or None,
        "motivo_consulta": visit.motivo or "Consulta medicina estética",
        "tratamiento": visit.tratamiento or None,
        "proxima_revision_fecha": None,
        "proxima_revision_descripcion": None,
        "observaciones": observaciones,
    }
    if existing:
        cur.execute(
            """
            UPDATE public.historial_clinico
            SET fecha = %(fecha)s,
                appointment_id = COALESCE(%(appointment_id)s, appointment_id),
                tipo = %(tipo)s,
                titulo = %(titulo)s,
                descripcion = %(descripcion)s,
                antecedentes_personales = %(antecedentes_personales)s,
                motivo_consulta = %(motivo_consulta)s,
                tratamiento = %(tratamiento)s,
                proxima_revision_fecha = %(proxima_revision_fecha)s,
                proxima_revision_descripcion = %(proxima_revision_descripcion)s,
                observaciones = %(observaciones)s
            WHERE id = %(id)s::uuid
            """,
            {**payload, "id": existing["id"]},
        )
        return existing["id"]

    cur.execute(
        """
        INSERT INTO public.historial_clinico (
          customer_id, company_id, fecha, appointment_id, tipo, titulo, descripcion,
          antecedentes_personales, motivo_consulta, tratamiento,
          proxima_revision_fecha, proxima_revision_descripcion, observaciones
        )
        VALUES (
          %(customer_id)s, %(company_id)s, %(fecha)s, %(appointment_id)s, %(tipo)s,
          %(titulo)s, %(descripcion)s, %(antecedentes_personales)s, %(motivo_consulta)s,
          %(tratamiento)s, %(proxima_revision_fecha)s, %(proxima_revision_descripcion)s,
          %(observaciones)s
        )
        RETURNING id::text
        """,
        payload,
    )
    return cur.fetchone()["id"]


def run_import(args: argparse.Namespace) -> dict[str, Any]:
    psycopg2, extras, _errors = _require_psycopg2()
    visits = parse_csv(args.csv)

    conn = psycopg2.connect(args.database_url)
    conn.autocommit = False
    report: dict[str, Any] = {
        "apply": args.apply,
        "csv": str(args.csv),
        "company_id": args.company_id,
        "patients_approx": len({v.normalized_name for v in visits if v.normalized_name}),
        "visits_total": len(visits),
        "planned_records": 0,
        "written_records": 0,
        "created_customers": 0,
        "birth_dates_set": 0,
        "linked_appointments": 0,
        "skipped_visits": [],
        "warnings": [],
        "sample_visits": [],
        "created_customer_names": [],
    }

    assignments: dict[int, str] = {}
    if args.assignments:
        assignments = load_assignments(args.assignments)
    report["assignments_loaded"] = len(assignments)

    used_appointments: set[str] = set()

    try:
        with conn.cursor(cursor_factory=extras.RealDictCursor) as cur:
            by_name, by_code, customers = load_customers(cur, args.company_id)
            report["customers_loaded"] = len(customers)

            for visit in visits:
                info = {
                    "line": visit.line,
                    "name": visit.name,
                    "fecha": visit.fecha or visit.raw_fecha,
                    "source_key": visit.source_key,
                    "is_first": visit.is_first,
                    "motivo": (visit.motivo or "")[:80],
                }
                if visit.issues and any(
                    i.startswith("missing_") or i.startswith("invalid_") for i in visit.issues
                ):
                    # Saltar solo si falta nombre o fecha
                    hard = [i for i in visit.issues if i in {"missing_customer_name", "missing_exam_date"} or i.startswith("invalid_date")]
                    if hard and not visit.fecha:
                        report["skipped_visits"].append({**info, "reason": hard[0], "issues": visit.issues})
                        continue

                assigned_code = assignments.get(visit.line)
                # Tambien probar por linea del bloque paciente (primera visita)
                if not assigned_code and visit.is_first:
                    assigned_code = assignments.get(visit.line)

                if assigned_code:
                    info["assigned_code"] = assigned_code
                    code_matches = by_code.get(assigned_code, [])
                    if len(code_matches) == 1:
                        customer, candidates, customer_issue = code_matches[0], [], None
                    elif not code_matches:
                        customer, candidates, customer_issue = None, [], "assigned_code_not_found"
                    else:
                        customer, candidates, customer_issue = None, code_matches, "assigned_code_ambiguous"
                else:
                    customer, candidates, customer_issue = find_customer(visit, by_name)

                if customer_issue or not customer:
                    # Si hay un unico candidato fuzzy, usarlo
                    if customer_issue == "customer_not_found" and candidates and len(candidates) == 1:
                        customer = candidates[0]
                        customer_issue = None
                    elif (
                        getattr(args, "create_customers", True)
                        and visit.name
                        and customer_issue == "customer_not_found"
                    ):
                        # Crear si no hay candidatos, o si los hay pero ninguno es suficientemente claro
                        # (ya se intento fuzzy arriba). Si hay candidatos, crear igual solo cuando
                        # el nombre no se parece demasiado a ninguno? Mejor: crear si no hay candidatos.
                        if len(candidates) == 0:
                            existing_created = by_name.get(visit.normalized_name, [])
                            if existing_created:
                                customer = existing_created[0]
                            elif args.apply:
                                customer = create_customer(cur, args.company_id, visit.name)
                                by_name.setdefault(normalize_name(customer["name"]), []).append(customer)
                                report["created_customer_names"].append(customer["name"])
                            else:
                                customer = {
                                    "id": f"new:{visit.normalized_name}",
                                    "company_id": args.company_id,
                                    "name": visit.name,
                                    "legacy_codcli": None,
                                    "birth_date": None,
                                }
                                by_name.setdefault(visit.normalized_name, []).append(customer)
                                report["created_customer_names"].append(visit.name)
                            customer_issue = None
                        else:
                            # Elegir el mejor candidato si ratio alto
                            best = max(
                                candidates,
                                key=lambda row: difflib.SequenceMatcher(
                                    None, visit.normalized_name, normalize_name(row["name"] or "")
                                ).ratio(),
                            )
                            best_ratio = difflib.SequenceMatcher(
                                None, visit.normalized_name, normalize_name(best["name"] or "")
                            ).ratio()
                            if best_ratio >= 0.84:
                                customer = best
                                customer_issue = None
                            else:
                                # Nombre del CSV no encaja claro → crear cliente nuevo
                                existing_created = by_name.get(visit.normalized_name, [])
                                if existing_created:
                                    customer = existing_created[0]
                                elif args.apply:
                                    customer = create_customer(cur, args.company_id, visit.name)
                                    by_name.setdefault(normalize_name(customer["name"]), []).append(customer)
                                    report["created_customer_names"].append(customer["name"])
                                else:
                                    customer = {
                                        "id": f"new:{visit.normalized_name}",
                                        "company_id": args.company_id,
                                        "name": visit.name,
                                        "legacy_codcli": None,
                                        "birth_date": None,
                                    }
                                    by_name.setdefault(visit.normalized_name, []).append(customer)
                                    report["created_customer_names"].append(visit.name)
                                customer_issue = None
                                report["warnings"].append(
                                    {
                                        **info,
                                        "reason": "created_despite_fuzzy_candidates",
                                        "best_ratio": round(best_ratio, 3),
                                        "candidates": [c["name"] for c in candidates[:5]],
                                    }
                                )
                    else:
                        report["skipped_visits"].append(
                            {
                                **info,
                                "reason": customer_issue or "customer_not_found",
                                "candidates": [
                                    {
                                        "id": item["id"],
                                        "name": item["name"],
                                        "legacy_codcli": item.get("legacy_codcli"),
                                    }
                                    for item in candidates
                                ],
                            }
                        )
                        continue

                if not customer:
                    report["skipped_visits"].append({**info, "reason": "customer_not_found"})
                    continue
                if not visit.fecha:
                    report["skipped_visits"].append({**info, "reason": "missing_or_invalid_exam_date"})
                    continue

                # En dry-run con cliente virtual no se consultan citas reales por id falso
                existing = None
                appointment_id = None
                if not str(customer["id"]).startswith("new:"):
                    existing = find_existing_history(cur, customer["id"], visit.source_key, visit.fecha)
                    appointment_id, appts, appt_issue = find_appointment(
                        cur, args.company_id, customer, visit.fecha
                    )
                    if appointment_id and (
                        appointment_id in used_appointments
                        or appointment_already_linked(
                            cur, appointment_id, existing["id"] if existing else None
                        )
                    ):
                        report["warnings"].append({**info, "reason": "appointment_already_linked"})
                        appointment_id = None
                    elif appt_issue:
                        report["warnings"].append(
                            {
                                **info,
                                "reason": appt_issue,
                                "appointments": [dict(item) for item in appts[:5]],
                            }
                        )
                    if appointment_id:
                        used_appointments.add(appointment_id)
                        report["linked_appointments"] += 1
                else:
                    report["warnings"].append({**info, "reason": "appointment_skipped_new_customer_dry_run"})

                report["planned_records"] += 1
                if len(report["sample_visits"]) < 12:
                    report["sample_visits"].append(
                        {
                            **info,
                            "customer_id": customer["id"],
                            "customer_name": customer["name"],
                            "appointment_id": appointment_id,
                            "antecedentes_preview": (visit.antecedentes or "")[:120],
                            "tratamiento_preview": (visit.tratamiento or "")[:120],
                        }
                    )

                if args.apply:
                    if visit.is_first and visit.edad:
                        if maybe_set_birth_date(cur, customer, visit.edad, visit.fecha):
                            report["birth_dates_set"] += 1
                    upsert_visit(cur, visit, args.company_id, customer, appointment_id)
                    report["written_records"] += 1

            if args.apply:
                conn.commit()
            else:
                conn.rollback()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    report["created_customer_names"] = sorted(set(report["created_customer_names"]))
    report["created_customers"] = len(report["created_customer_names"])
    return report


def main() -> int:
    load_dotenv(Path(".env"))
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument(
        "--company-id",
        default=os.environ.get("PROMOTE_COMPANY_ID")
        or os.environ.get("LEGACY_COMPANY_ID")
        or DEFAULT_IMPORT_COMPANY_ID,
    )
    parser.add_argument("--database-url", default=os.environ.get("SUPABASE_DB_URL"))
    parser.add_argument("--report", type=Path, default=Path("tmp/medical_aesthetic_import_report.json"))
    parser.add_argument(
        "--assignments",
        type=Path,
        default=None,
        help="Excel con asignaciones manuales (Linea CSV -> codigo de cliente).",
    )
    parser.add_argument("--apply", action="store_true", help="Escribe datos. Sin este flag solo dry-run.")
    args = parser.parse_args()

    if not args.csv.exists():
        print(f"No existe el CSV: {args.csv}", file=sys.stderr)
        return 2
    if args.assignments and not args.assignments.exists():
        print(f"No existe el Excel de asignaciones: {args.assignments}", file=sys.stderr)
        return 2
    if not args.company_id:
        print("Falta --company-id", file=sys.stderr)
        return 2
    if not args.database_url:
        print("Falta --database-url o SUPABASE_DB_URL en .env", file=sys.stderr)
        return 2

    report = run_import(args)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(
        "apply" if args.apply else "dry-run",
        f"pacientes≈{report['patients_approx']}",
        f"visitas={report['visits_total']}",
        f"planificadas={report['planned_records']}",
        f"escritas={report['written_records']}",
        f"clientes_nuevos={report.get('created_customers', 0)}",
        f"citas_linkeadas={report['linked_appointments']}",
        f"saltadas={len(report['skipped_visits'])}",
        f"avisos={len(report['warnings'])}",
        f"fn_seteadas={report['birth_dates_set']}",
        f"reporte={args.report}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
