#!/usr/bin/env python3
"""Genera un Excel con los pacientes del CSV de medicina estetica que no se
pudieron identificar como cliente, para resolverlos a mano aportando el codigo.

Lee el informe JSON producido por import_medical_aesthetic_history.py.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from import_medical_aesthetic_history import read_text_guess  # noqa: E402


CUSTOMER_REASONS = {
    "ambiguous_customer": "Varios clientes con el mismo nombre",
    "customer_not_found": "No se encontro cliente",
    "missing_customer_name": "Falta nombre de paciente",
    "missing_or_invalid_exam_date": "Fecha de examen invalida",
}

DEFAULT_REPORT = Path("tmp/medical_aesthetic_import_apply_report.json")
DEFAULT_CSV = Path(r"C:\Users\OportoW11\Desktop\Medicina Estética\Fichas medicina.csv")
DEFAULT_OUTPUT = Path(
    r"C:\Users\OportoW11\Desktop\Medicina Estética\pacientes_sin_identificar.xlsx"
)


def load_csv_lines(path: Path) -> list[str]:
    if not path.exists():
        return []
    return read_text_guess(path).splitlines()


def block_context(lines: list[str], start_line: int, max_lines: int = 6) -> str:
    if not lines or not start_line:
        return ""
    idx = start_line - 1
    collected: list[str] = []
    for raw in lines[idx : idx + 18]:
        text = raw.strip()
        if not text:
            continue
        collected.append(text)
        if len(collected) >= max_lines:
            break
    return "\n".join(collected)


def candidate_text(candidates: list[dict]) -> str:
    parts = []
    for item in candidates:
        name = item.get("name") or "(sin nombre)"
        code = item.get("legacy_codcli")
        parts.append(f"{name} -> {code}" if code else f"{name} -> (sin codigo)")
    return "\n".join(parts)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.report.exists():
        print(f"No existe el informe: {args.report}")
        return 2

    data = json.loads(args.report.read_text(encoding="utf-8"))
    skipped = data.get("skipped_blocks", [])
    csv_lines = load_csv_lines(args.csv)

    seen: set[tuple] = set()
    rows = []
    for block in skipped:
        reason = block.get("reason", "")
        if reason not in CUSTOMER_REASONS:
            continue
        key = (block.get("line"), block.get("source_key"))
        if key in seen:
            continue
        seen.add(key)
        rows.append(block)

    rows.sort(key=lambda b: (b.get("name") or "").lower())

    wb = Workbook()
    ws = wb.active
    ws.title = "Sin identificar"

    headers = [
        "Linea CSV",
        "Nombre en CSV",
        "Fecha examen",
        "Motivo",
        "Candidatos encontrados (nombre -> codigo)",
        "Contexto del CSV",
        "CODIGO CLIENTE (rellenar)",
        "Notas",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    fill_action = PatternFill("solid", fgColor="FFF2CC")
    for block in rows:
        candidates = block.get("candidates", [])
        ws.append(
            [
                block.get("line"),
                block.get("name"),
                block.get("fecha"),
                CUSTOMER_REASONS.get(block.get("reason", ""), block.get("reason", "")),
                candidate_text(candidates),
                block_context(csv_lines, block.get("line") or 0),
                "",
                "",
            ]
        )
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=7).fill = fill_action
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    widths = [10, 32, 14, 32, 50, 55, 24, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Excel generado: {args.output} ({len(rows)} pacientes sin identificar)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
